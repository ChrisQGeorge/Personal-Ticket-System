import csv
import io
from datetime import datetime

from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..database import get_db
from ..models import Profile, Ticket, User

router = APIRouter(tags=["import"])

VALID_PRIORITIES = {"very low", "low", "default", "high", "very high"}
VALID_STATUSES = {"open", "in-progress", "completed", "skipped"}
MAX_IMPORT_ROWS = 5000


def parse_date(val):
    """Try to parse a date from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    # openpyxl may return a date object directly
    from datetime import date as date_type

    if isinstance(val, date_type):
        return val
    val = str(val).strip()
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def parse_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if not val_str:
        return None
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return None


def normalize_headers(headers):
    """Normalize header names to lowercase, stripped."""
    return [str(h).strip().lower() if h else "" for h in headers]


FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r', '\n')

def _sanitize_cell(value: str) -> str:
    """Strip formula prefixes to prevent CSV/Excel injection."""
    if value and isinstance(value, str):
        while value and value[0] in FORMULA_PREFIXES:
            value = value[1:]
    return value.strip()


def _get_cell(row, col_map, *names):
    """Get a cell value by trying multiple column name variants."""
    for name in names:
        if name in col_map and col_map[name] < len(row):
            val = row[col_map[name]]
            return val
    return None


def rows_to_tickets(rows, headers, db: Session, profile_id: Optional[int] = None):
    """Parse rows into tickets. Returns (imported_count, errors)."""
    col_map = {h: i for i, h in enumerate(normalize_headers(headers))}
    imported = 0
    errors = []

    for row_num, row in enumerate(rows, start=2):  # row 1 is headers
        try:
            title_val = _get_cell(row, col_map, "title")
            title = _sanitize_cell(str(title_val)) if title_val else ""
            if not title or title.lower() == "none":
                continue

            description_val = _get_cell(row, col_map, "description")
            description = _sanitize_cell(str(description_val)) if description_val else None
            if description and description.lower() == "none":
                description = None

            priority_val = _get_cell(row, col_map, "priority")
            priority_str = str(priority_val).strip().lower() if priority_val else "default"
            if not priority_str or priority_str == "none" or priority_str not in VALID_PRIORITIES:
                priority_str = "default"

            status_val = _get_cell(row, col_map, "status")
            status_str = str(status_val).strip().lower() if status_val else "open"
            if not status_str or status_str == "none" or status_str not in VALID_STATUSES:
                status_str = "open"

            due_date = parse_date(_get_cell(row, col_map, "due_date", "due date"))

            est_hours = parse_float(_get_cell(row, col_map, "est_hours", "est hours"))

            ticket = Ticket(
                title=title,
                description=description,
                priority=priority_str,
                status=status_str,
                due_date=due_date,
                est_hours=est_hours,
                profile_id=profile_id,
            )
            db.add(ticket)
            imported += 1
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    db.commit()
    return imported, errors


@router.post("/import")
async def import_tickets(
    file: UploadFile = File(...),
    profile_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Import tickets from a CSV or Excel file."""
    # Verify profile ownership if profile_id provided
    if profile_id is not None:
        profile = db.query(Profile).filter(Profile.id == profile_id, Profile.user_id == user.id).first()
        if not profile:
            raise HTTPException(status_code=403, detail="Profile does not belong to you")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided.")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB.")

    if ext == "csv":
        try:
            text = contents.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = contents.decode("latin-1")

        # Auto-detect delimiter using csv.Sniffer
        try:
            dialect = csv.Sniffer().sniff(text[:4096])
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="CSV file is empty.")

        headers = all_rows[0]
        data_rows = all_rows[1:]
        if len(data_rows) > MAX_IMPORT_ROWS:
            raise HTTPException(400, f"Too many rows. Maximum is {MAX_IMPORT_ROWS}.")
        imported, errors = rows_to_tickets(data_rows, headers, db, profile_id=profile_id)

    elif ext in ("xlsx", "xls"):
        try:
            wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            headers = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="Excel file is empty.")

        data_rows = list(rows_iter)
        wb.close()
        if len(data_rows) > MAX_IMPORT_ROWS:
            raise HTTPException(400, f"Too many rows. Maximum is {MAX_IMPORT_ROWS}.")
        imported, errors = rows_to_tickets(data_rows, list(headers), db, profile_id=profile_id)

    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '.{ext}'. Please upload a .csv or .xlsx file.",
        )

    return {"imported": imported, "errors": errors}


@router.get("/import/template")
def download_template(user: User = Depends(get_current_user)):
    """Download an Excel template for importing tickets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = ["Title", "Description", "Priority", "Due Date", "Est Hours", "Status"]
    ws.append(headers)
    # Example row
    ws.append(["Example Task", "Description of the task", "default", "2025-12-31", 1.5, "open"])

    # Style header row bold
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pts_import_template.xlsx"},
    )
